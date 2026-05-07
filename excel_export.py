"""
Excel export for 3LC Conference Finder results.
Generates a formatted .xlsx file with the 6 required columns.
"""

import io
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


COLUMNS = [
    "Conference Name",
    "Location",
    "Date",
    "Industry",
    "COI (Companies of Interest)",
    "Conference Size"
]

COLUMN_WIDTHS = [40, 25, 20, 25, 60, 20]

# 3LC brand-ish colors
HEADER_BG = "1A1A2E"       # Dark navy
HEADER_FONT = "FFFFFF"     # White
ROW_BG_ODD = "F0F4FF"      # Light blue-gray
ROW_BG_EVEN = "FFFFFF"     # White
ACCENT = "0066CC"          # Blue accent


def generate_excel(conferences: list[dict], filters: dict | None = None) -> bytes:
    """
    Generate a formatted Excel file from conference data.

    Args:
        conferences: List of conference dicts
        filters: Optional dict of search filters used (for the info sheet)

    Returns:
        Bytes of the .xlsx file
    """
    wb = openpyxl.Workbook()

    # --- Main data sheet ---
    ws = wb.active
    ws.title = "Conferences"

    _write_header(ws)
    _write_data(ws, conferences)
    _apply_formatting(ws, len(conferences))

    # --- Info sheet ---
    if filters:
        ws_info = wb.create_sheet("Search Parameters")
        _write_info_sheet(ws_info, filters, len(conferences))

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _write_header(ws):
    """Write and style the header row."""
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color=HEADER_FONT, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        bottom=Side(style="medium", color=ACCENT)
    )

    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, COLUMN_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 30


def _write_data(ws, conferences: list[dict]):
    """Write conference data rows."""
    field_map = [
        "conference_name",
        "location",
        "date",
        "industry",
        "companies_of_interest",
        "conference_size"
    ]

    for row_idx, conf in enumerate(conferences, start=2):
        is_odd = (row_idx % 2 != 0)
        row_bg = ROW_BG_ODD if is_odd else ROW_BG_EVEN

        row_fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
        row_font = Font(name="Calibri", size=10)
        wrap_align = Alignment(vertical="top", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

        for col_idx, field in enumerate(field_map, start=1):
            value = conf.get(field, "Unknown") or "Unknown"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = row_font

            # Center-align location, date, and size columns
            if col_idx in (2, 3, 6):
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align

        ws.row_dimensions[row_idx].height = 60


def _apply_formatting(ws, num_rows: int):
    """Apply table formatting and freeze header."""
    ws.freeze_panes = "A2"

    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{num_rows + 1}"

    # Title above header
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value="3LC.ai — Conference Research Report")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=ACCENT)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    generated_cell = ws.cell(
        row=1, column=len(COLUMNS),
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
    )
    generated_cell.font = Font(name="Calibri", size=9, color="888888")
    generated_cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[1].height = 25
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS) - 1)}1")


def _write_info_sheet(ws, filters: dict, num_results: int):
    """Write search parameters to a separate sheet."""
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    header_font = Font(name="Calibri", bold=True, size=12, color=ACCENT)
    label_font = Font(name="Calibri", bold=True, size=10)
    value_font = Font(name="Calibri", size=10)

    ws.cell(row=1, column=1, value="Search Parameters").font = header_font
    ws.cell(row=2, column=1, value="Parameter").font = label_font
    ws.cell(row=2, column=2, value="Value").font = label_font

    params = [
        ("Industries", ", ".join(filters.get("industries", []))),
        ("Date Range", filters.get("date_range", "N/A")),
        ("Min Conference Size", f"{filters.get('min_size', 0)}+ companies"),
        ("Regions", ", ".join(filters.get("regions", ["All US"]))),
        ("Results Found", str(num_results)),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    for row_idx, (param, value) in enumerate(params, start=3):
        ws.cell(row=row_idx, column=1, value=param).font = label_font
        ws.cell(row=row_idx, column=2, value=value).font = value_font
